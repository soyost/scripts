#!/usr/bin/env python3

"""Validate Nexus interfaces and BGP neighbors for a carrier circuit.

Example:
    python bgp_verify.py --circuit 447550116 --xlsx circuits.xlsx

The circuit workbook must contain these logical columns (matching is flexible):
    Service/Circuit
    "A"-END-Router
    "Z"-END-Router

Router cells are expected to resemble:
    NXWASEADX01 - eth1/46
"""

from __future__ import annotations

import argparse
import csv
import ipaddress
import re
import sys
from dataclasses import dataclass, asdict
from datetime import datetime
from getpass import getpass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from netmiko import ConnectHandler
from netmiko.exceptions import (
    NetmikoAuthenticationException,
    NetmikoTimeoutException,
)
from paramiko.ssh_exception import SSHException


DEFAULT_CIRCUIT_FILE = "circuits.xlsx"
DEFAULT_RESULTS_FILE = "bgp-verification-results.csv"
DEFAULT_VRFS = ("OAM", "PROD")
DEVICE_TYPES = ("cisco_nxos", "cisco_xe", "cisco_ios")
CIRCUIT_RE = re.compile(r"^4\d{8}$")


@dataclass(frozen=True)
class Endpoint:
    side: str
    host: str
    parent_interface: str
    raw_value: str


@dataclass
class VerificationRow:
    timestamp: str
    circuit_id: str
    side: str
    host: str
    parent_interface: str
    csv_router_value: str
    connection_status: str = ""
    device_type: str = ""
    circuit_config_verified: str = "NO"
    circuit_config_proof: str = ""
    vrf: str = ""
    subinterface: str = ""
    interface_ip: str = ""
    interface_status: str = ""
    interface_healthy: str = "NO"
    neighbor_ip: str = ""
    remote_as: str = ""
    bgp_state: str = ""
    bgp_uptime: str = ""
    prefixes_received: str = ""
    in_queue: str = ""
    out_queue: str = ""
    reciprocal_peer_verified: str = "NO"
    result: str = ""
    error: str = ""


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split())


def normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.lower())


def find_column(fieldnames: Iterable[str], aliases: Iterable[str]) -> str:
    normalized = {normalize_header(name): name for name in fieldnames if name}
    for alias in aliases:
        match = normalized.get(normalize_header(alias))
        if match:
            return match
    raise KeyError(f"Could not find a CSV column matching: {', '.join(aliases)}")


def validate_circuit_id(value: str) -> str:
    circuit_id = value.strip()
    if not CIRCUIT_RE.fullmatch(circuit_id):
        raise ValueError("Circuit ID must be 9 digits and begin with 4.")
    return circuit_id


def normalize_interface_name(name: str) -> str:
    value = name.strip()
    replacements = {
        "ethernet": "Eth",
        "eth": "Eth",
        "port-channel": "Po",
        "portchannel": "Po",
        "po": "Po",
    }
    lower = value.lower()
    for prefix, replacement in replacements.items():
        if lower.startswith(prefix):
            return replacement + value[len(prefix):]
    return value


def parent_interface(name: str) -> str:
    return normalize_interface_name(name).split(".", 1)[0]


def parse_router_cell(side: str, raw_value: str) -> Endpoint:
    raw = clean_text(raw_value)
    if not raw or raw.lower() == "unknown":
        raise ValueError(f'{side}-end router is missing or "Unknown" in the workbook.')

    # Expected examples:
    #   NXWASEADX01 - eth1/46
    #   NXKSOPHEX01-eth1/46
    match = re.search(
        r"(?P<host>[A-Za-z0-9_.-]+)\s*-\s*(?P<intf>(?:Eth(?:ernet)?|Po(?:rt-?channel)?)\s*\d+(?:/\d+)*(?:\.\d+)?)",
        raw,
        re.IGNORECASE,
    )
    if not match:
        raise ValueError(
            f"Could not parse {side}-end router/interface from workbook value: {raw!r}"
        )

    return Endpoint(
        side=side,
        host=match.group("host").upper(),
        parent_interface=parent_interface(match.group("intf").replace(" ", "")),
        raw_value=raw,
    )


def cell_text(value: object) -> str:
    """Convert an Excel cell value to predictable text."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean_text(value)


def load_circuit(xlsx_path: Path, circuit_id: str) -> tuple[dict[str, str], list[Endpoint]]:
    """Find a circuit in an Excel workbook and return its two endpoints.

    Every worksheet is searched. The header row may appear anywhere in a sheet.
    Extra notes, formatting, and non-circuit rows are ignored.
    """
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)

    circuit_aliases = (
        "Service/Circuit",
        "Service Circuit",
        "Circuit",
        "Circuit ID",
    )
    a_router_aliases = (
        '"A"-END-Router',
        "A-END-Router",
        "A End Router",
        "A Router",
    )
    z_router_aliases = (
        '"Z"-END-Router',
        "Z-END-Router",
        "Z End Router",
        "Z Router",
    )

    try:
        for worksheet in workbook.worksheets:
            header_map: dict[str, int] | None = None

            for row in worksheet.iter_rows(values_only=True):
                values = [cell_text(value) for value in row]

                if header_map is None:
                    normalized = {
                        normalize_header(value): index
                        for index, value in enumerate(values)
                        if value
                    }

                    def find_index(aliases: Iterable[str]) -> int | None:
                        for alias in aliases:
                            index = normalized.get(normalize_header(alias))
                            if index is not None:
                                return index
                        return None

                    circuit_index = find_index(circuit_aliases)
                    a_router_index = find_index(a_router_aliases)
                    z_router_index = find_index(z_router_aliases)

                    if None not in (circuit_index, a_router_index, z_router_index):
                        header_map = {
                            "circuit": int(circuit_index),
                            "a_router": int(a_router_index),
                            "z_router": int(z_router_index),
                        }
                    continue

                circuit_value = values[header_map["circuit"]] if header_map["circuit"] < len(values) else ""
                row_circuit = re.sub(r"\D", "", circuit_value)

                # This also ignores notes and status rows below the circuit table.
                if row_circuit != circuit_id:
                    continue

                a_value = values[header_map["a_router"]] if header_map["a_router"] < len(values) else ""
                z_value = values[header_map["z_router"]] if header_map["z_router"] < len(values) else ""

                record = {
                    "worksheet": worksheet.title,
                    "service_circuit": circuit_value,
                    "a_router": a_value,
                    "z_router": z_value,
                }
                endpoints = [
                    parse_router_cell("A", a_value),
                    parse_router_cell("Z", z_value),
                ]
                return record, endpoints
    finally:
        workbook.close()

    raise LookupError(f"Circuit {circuit_id} was not found in {xlsx_path}.")


def connect_device(host: str, username: str, password: str):
    last_error = ""
    for device_type in DEVICE_TYPES:
        conn = None
        try:
            conn = ConnectHandler(
                host=host,
                username=username,
                password=password,
                device_type=device_type,
                fast_cli=False,
                conn_timeout=12,
                banner_timeout=15,
                auth_timeout=15,
            )
            prompt = conn.find_prompt()
            if not prompt:
                raise RuntimeError("Connected, but no device prompt was detected.")
            return conn, device_type
        except (
            NetmikoTimeoutException,
            NetmikoAuthenticationException,
            SSHException,
            OSError,
            RuntimeError,
        ) as exc:
            last_error = clean_text(exc)
            if conn:
                try:
                    conn.disconnect()
                except Exception:
                    pass

    raise ConnectionError(last_error or "Unable to connect or detect platform.")


def parse_ip_interface_brief(output: str) -> list[dict[str, str]]:
    results: list[dict[str, str]] = []
    for line in output.splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("Interface"):
            continue

        parts = stripped.split()
        if len(parts) < 3:
            continue

        interface = normalize_interface_name(parts[0])
        ip_address = parts[1]
        status = " ".join(parts[2:])

        if not re.match(r"^(?:Eth|Po)\d", interface, re.IGNORECASE):
            continue

        results.append(
            {
                "interface": interface,
                "ip_address": ip_address,
                "status": status,
            }
        )
    return results


def parse_interface_config(output: str) -> tuple[str, str]:
    ip_cidr = ""
    vrf = ""

    if match := re.search(r"^\s*ip address\s+(\d+\.\d+\.\d+\.\d+/\d+)", output, re.MULTILINE):
        ip_cidr = match.group(1)

    if match := re.search(r"^\s*vrf member\s+(\S+)", output, re.MULTILINE):
        vrf = match.group(1)

    return ip_cidr, vrf


def interface_is_healthy(status: str) -> bool:
    normalized = status.lower()
    return all(token in normalized for token in ("protocol-up", "link-up", "admin-up"))


def parse_bgp_summary(output: str) -> list[dict[str, str]]:
    neighbors: list[dict[str, str]] = []
    in_table = False

    for line in output.splitlines():
        if re.search(r"^Neighbor\s+V\s+AS\s+", line):
            in_table = True
            continue
        if not in_table:
            continue

        stripped = line.strip()
        if not stripped:
            continue

        parts = stripped.split()
        if len(parts) < 10 or not re.fullmatch(r"\d+\.\d+\.\d+\.\d+", parts[0]):
            continue

        state_or_prefixes = parts[-1]
        established = state_or_prefixes.isdigit()
        neighbors.append(
            {
                "neighbor_ip": parts[0],
                "version": parts[1],
                "remote_as": parts[2],
                "msg_received": parts[3],
                "msg_sent": parts[4],
                "table_version": parts[5],
                "in_queue": parts[6],
                "out_queue": parts[7],
                "uptime": parts[8],
                "state": "Established" if established else state_or_prefixes,
                "prefixes": state_or_prefixes if established else "",
            }
        )

    return neighbors


def find_neighbor_for_interface(
    local_cidr: str,
    neighbors: list[dict[str, str]],
) -> dict[str, str] | None:
    local = ipaddress.ip_interface(local_cidr)
    matches = []
    for neighbor in neighbors:
        address = ipaddress.ip_address(neighbor["neighbor_ip"])
        if address in local.network and address != local.ip:
            matches.append(neighbor)

    if len(matches) == 1:
        return matches[0]
    return None


def verify_endpoint(
    endpoint: Endpoint,
    circuit_id: str,
    username: str,
    password: str,
    vrfs: tuple[str, ...],
) -> list[VerificationRow]:
    timestamp = datetime.now().isoformat(timespec="seconds")
    base = VerificationRow(
        timestamp=timestamp,
        circuit_id=circuit_id,
        side=endpoint.side,
        host=endpoint.host,
        parent_interface=endpoint.parent_interface,
        csv_router_value=endpoint.raw_value,
    )

    conn = None
    try:
        print(f"{endpoint.host}: connecting")
        conn, device_type = connect_device(endpoint.host, username, password)
        base.connection_status = "CONNECTED"
        base.device_type = device_type

        circuit_output = conn.send_command(
            f"show running-config | include {circuit_id}",
            read_timeout=20,
        )
        proof_lines = [
            clean_text(line)
            for line in circuit_output.splitlines()
            if circuit_id in line
        ]
        base.circuit_config_verified = "YES" if proof_lines else "NO"
        base.circuit_config_proof = " | ".join(proof_lines)

        found_rows: list[VerificationRow] = []

        for vrf in vrfs:
            brief_output = conn.send_command(
                f"show ip interface brief vrf {vrf}",
                read_timeout=20,
            )
            matching_interfaces = [
                item
                for item in parse_ip_interface_brief(brief_output)
                if parent_interface(item["interface"]).lower()
                == endpoint.parent_interface.lower()
                and "." in item["interface"]
            ]

            if not matching_interfaces:
                continue

            bgp_output = conn.send_command(
                f"show bgp ipv4 unicast summary vrf {vrf}",
                read_timeout=30,
            )
            bgp_neighbors = parse_bgp_summary(bgp_output)

            for interface_data in matching_interfaces:
                row = VerificationRow(**asdict(base))
                row.vrf = vrf
                row.subinterface = interface_data["interface"]
                row.interface_status = interface_data["status"]
                row.interface_healthy = (
                    "YES" if interface_is_healthy(interface_data["status"]) else "NO"
                )

                config_output = conn.send_command(
                    f"show running-config interface {interface_data['interface']}",
                    read_timeout=20,
                )
                ip_cidr, config_vrf = parse_interface_config(config_output)
                row.interface_ip = ip_cidr or interface_data["ip_address"]

                if config_vrf and config_vrf.upper() != vrf.upper():
                    row.error = (
                        f"Interface config says VRF {config_vrf}, but it was found "
                        f"under VRF {vrf}."
                    )

                if not ip_cidr:
                    row.result = "UNKNOWN"
                    row.error = clean_text(
                        f"{row.error} Could not determine interface prefix length."
                    )
                    found_rows.append(row)
                    continue

                neighbor = find_neighbor_for_interface(ip_cidr, bgp_neighbors)
                if not neighbor:
                    row.result = "FAIL" if row.interface_healthy == "YES" else "FAIL"
                    row.error = clean_text(
                        f"{row.error} No unique BGP neighbor in {ipaddress.ip_interface(ip_cidr).network}."
                    )
                    found_rows.append(row)
                    continue

                row.neighbor_ip = neighbor["neighbor_ip"]
                row.remote_as = neighbor["remote_as"]
                row.bgp_state = neighbor["state"]
                row.bgp_uptime = neighbor["uptime"]
                row.prefixes_received = neighbor["prefixes"]
                row.in_queue = neighbor["in_queue"]
                row.out_queue = neighbor["out_queue"]

                if row.interface_healthy == "YES" and row.bgp_state == "Established":
                    row.result = "PASS"
                else:
                    row.result = "FAIL"

                found_rows.append(row)

        if not found_rows:
            row = VerificationRow(**asdict(base))
            row.result = "UNKNOWN"
            row.error = (
                f"No subinterfaces under {endpoint.parent_interface} were found in "
                f"VRFs {', '.join(vrfs)}."
            )
            return [row]

        return found_rows

    except Exception as exc:
        base.connection_status = base.connection_status or "FAILED"
        base.result = "ERROR"
        base.error = clean_text(exc)
        return [base]
    finally:
        if conn:
            try:
                conn.disconnect()
            except Exception:
                pass


def mark_reciprocal_peers(rows: list[VerificationRow]) -> None:
    for row in rows:
        if not row.neighbor_ip or not row.vrf:
            continue

        for candidate in rows:
            if candidate.host == row.host or candidate.vrf.upper() != row.vrf.upper():
                continue
            try:
                candidate_local_ip = str(ipaddress.ip_interface(candidate.interface_ip).ip)
            except ValueError:
                continue

            if candidate_local_ip == row.neighbor_ip:
                row.reciprocal_peer_verified = "YES"
                break


def print_report(circuit_id: str, endpoints: list[Endpoint], rows: list[VerificationRow]) -> None:
    print("\n" + "=" * 78)
    print(f"Circuit {circuit_id} BGP verification")
    print("=" * 78)

    for endpoint in endpoints:
        print(f"\n{endpoint.side}-END: {endpoint.host} {endpoint.parent_interface}")
        endpoint_rows = [row for row in rows if row.side == endpoint.side]

        for row in endpoint_rows:
            if not row.vrf:
                print(f"  Result: {row.result}")
                if row.error:
                    print(f"  Error: {row.error}")
                continue

            print(f"  VRF {row.vrf} / {row.subinterface}")
            print(f"    Circuit in config: {row.circuit_config_verified}")
            print(f"    Interface: {row.interface_status} ({row.interface_healthy})")
            print(f"    Local IP: {row.interface_ip or 'UNKNOWN'}")
            print(f"    Neighbor: {row.neighbor_ip or 'NOT FOUND'}")
            print(f"    Remote AS: {row.remote_as or 'UNKNOWN'}")
            print(f"    BGP state: {row.bgp_state or 'NOT FOUND'}")
            print(f"    BGP uptime: {row.bgp_uptime or 'N/A'}")
            print(f"    Prefixes received: {row.prefixes_received or 'N/A'}")
            print(f"    InQ/OutQ: {row.in_queue or 'N/A'}/{row.out_queue or 'N/A'}")
            print(f"    Peer IP Validation: {row.reciprocal_peer_verified}")
            print(f"    Result: {row.result}")
            if row.error:
                print(f"    Note: {row.error}")

    meaningful_rows = [row for row in rows if row.vrf]
    if meaningful_rows and all(row.result == "PASS" for row in meaningful_rows):
        overall = "PASS"
    elif any(row.result in {"FAIL", "ERROR"} for row in rows):
        overall = "FAIL"
    else:
        overall = "UNKNOWN"

    print("\n" + "-" * 78)
    print(f"OVERALL RESULT: {overall}")
    print("-" * 78)


def write_results(path: Path, rows: list[VerificationRow]) -> None:
    fieldnames = list(VerificationRow.__dataclass_fields__.keys())
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(asdict(row))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Validate Nexus interfaces and BGP neighbors by carrier circuit ID."
    )
    parser.add_argument(
        "--circuit",
        help="Nine-digit circuit ID beginning with 4, for example 447550116.",
    )
    parser.add_argument(
        "--xlsx",
        default=DEFAULT_CIRCUIT_FILE,
        help=f"Circuit Excel workbook path (default: {DEFAULT_CIRCUIT_FILE}).",
    )
    parser.add_argument(
        "--output",
        default=DEFAULT_RESULTS_FILE,
        help=f"Results CSV path (default: {DEFAULT_RESULTS_FILE}).",
    )
    parser.add_argument(
        "--vrfs",
        nargs="+",
        default=list(DEFAULT_VRFS),
        help="VRFs to inspect (default: OAM PROD).",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    try:
        circuit_id = validate_circuit_id(
            args.circuit or input("Circuit ID: ").strip()
        )
        xlsx_path = Path(args.xlsx)
        if not xlsx_path.is_file():
            raise FileNotFoundError(f"Circuit workbook not found: {xlsx_path}")

        circuit_record, endpoints = load_circuit(xlsx_path, circuit_id)
    except (ValueError, KeyError, LookupError, FileNotFoundError) as exc:
        print(f"ERROR: {exc}")
        return 2

    print(f"Circuit {circuit_id} found in {xlsx_path} (sheet: {circuit_record['worksheet']})")
    for endpoint in endpoints:
        print(f"  {endpoint.side}-END: {endpoint.host} {endpoint.parent_interface}")

    username = input("TACACS username: ").strip()
    password = getpass("TACACS password: ")
    if not username or not password:
        print("ERROR: TACACS username and password are required.")
        return 2

    rows: list[VerificationRow] = []
    for endpoint in endpoints:
        rows.extend(
            verify_endpoint(
                endpoint=endpoint,
                circuit_id=circuit_id,
                username=username,
                password=password,
                vrfs=tuple(args.vrfs),
            )
        )

    mark_reciprocal_peers(rows)
    print_report(circuit_id, endpoints, rows)

    output_path = Path(args.output)
    write_results(output_path, rows)
    print(f"\nResults written to: {output_path}")

    return 0 if all(row.result == "PASS" for row in rows if row.vrf) else 1


if __name__ == "__main__":
    sys.exit(main())