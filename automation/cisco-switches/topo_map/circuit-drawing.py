from pathlib import Path
import re

from openpyxl import load_workbook
from pyvis.network import Network


INPUT_FILE = Path("circuits.xlsx")
OUTPUT_FILE = Path("circuits-topology.html")

# Your spreadsheet has a blank first row.
HEADER_ROW = 2


def clean(value):
    """Convert Excel values to clean strings."""
    if value is None:
        return ""
    return str(value).strip()


def parse_endpoint(value):
    """
    Parse values such as:

        NXKSOPHEX01 - eth1/52
        NXWASEADW01 -eth1/47
        NXWASEADX01 - eth1/46 (per lldp)
        lag-32-3890-99.ear2.Seattle1.Level3.net
        Unknown

    Returns:
        device, interface
    """

    value = clean(value)

    if not value:
        return None, None

    # Remove informational annotation from the interface/device string.
    value = re.sub(
        r"\s*\(per\s+lldp\)\s*$",
        "",
        value,
        flags=re.IGNORECASE,
    ).strip()

    # Nexus names in this spreadsheet begin with NX.
    # Capture:
    #   NXKSOPHEX01 - eth1/52
    #   NXWASEADW01 -eth1/47
    match = re.match(
        r"^(NX[A-Za-z0-9_-]+)\s*-\s*(.+)$",
        value,
        flags=re.IGNORECASE,
    )

    if match:
        device = match.group(1).upper()
        interface = match.group(2).strip()
        return device, interface

    # Anything else is an external/provider endpoint.
    return value, None


def read_circuits(filename):
    workbook = load_workbook(filename, data_only=True)
    worksheet = workbook.active

    # Build dictionary of:
    # header name -> column number
    headers = {}

    for cell in worksheet[HEADER_ROW]:
        if cell.value:
            headers[clean(cell.value)] = cell.column

    print("Detected spreadsheet columns:")

    for header in headers:
        print(f"  {header}")

    required = {
        "Service/Circuit",
        "Service Component ID",
        "Type",
        "A-End - City/State",
        "A-End - Address",
        '"A" -END -Router',
        "Z-End - Address",
        '"Z"-END-Router',
        "Notes",
    }

    missing = required - set(headers)

    if missing:
        raise ValueError(
            "Spreadsheet is missing required columns: "
            + ", ".join(sorted(missing))
        )

    circuits = []

    for row_num in range(HEADER_ROW + 1, worksheet.max_row + 1):

        def get(column_name):
            column = headers[column_name]
            return clean(worksheet.cell(row=row_num, column=column).value)

        circuit_number = get("Service/Circuit")

        # Ignore NX List / Present rows at the bottom of your workbook.
        if not circuit_number:
            continue

        a_raw = get('"A" -END -Router')
        z_raw = get('"Z"-END-Router')

        a_device, a_interface = parse_endpoint(a_raw)
        z_device, z_interface = parse_endpoint(z_raw)

        # A circuit with neither endpoint isn't drawable.
        if not a_device and not z_device:
            print(
                f"Skipping circuit {circuit_number}: "
                "no A or Z router information"
            )
            continue

        circuits.append(
            {
                "circuit": circuit_number,
                "component": get("Service Component ID"),
                "type": get("Type"),
                "city": get("A-End - City/State"),
                "a_address": get("A-End - Address"),
                "a_raw": a_raw,
                "a_device": a_device,
                "a_interface": a_interface,
                "z_address": get("Z-End - Address"),
                "z_raw": z_raw,
                "z_device": z_device,
                "z_interface": z_interface,
                "notes": get("Notes"),
            }
        )

    return circuits


def is_nexus(device):
    """Identify Nexus devices from your naming convention."""
    if not device:
        return False

    return device.upper().startswith("NX")


def add_node(network, node_id, label=None, external=False):
    """
    Add a node only if it doesn't already exist.
    """

    existing_nodes = {node["id"] for node in network.nodes}

    if node_id in existing_nodes:
        return

    if external:
        network.add_node(
            node_id,
            label=label or node_id,
            shape="ellipse",
            title=f"External endpoint<br>{node_id}",
            font={"size": 14},
        )
    else:
        network.add_node(
            node_id,
            label=label or node_id,
            shape="box",
            title=f"Nexus switch<br>{node_id}",
            font={"size": 16},
            margin=12,
        )


def build_hover_text(circuit):
    """
    Information displayed when hovering over a circuit.
    """

    lines = [
        f"<b>Circuit:</b> {circuit['circuit']}",
    ]

    if circuit["component"]:
        lines.append(
            f"<b>Service Component:</b> {circuit['component']}"
        )

    if circuit["type"]:
        lines.append(
            f"<b>Type:</b> {circuit['type']}"
        )

    if circuit["a_raw"]:
        lines.append(
            f"<b>A End:</b> {circuit['a_raw']}"
        )

    if circuit["z_raw"]:
        lines.append(
            f"<b>Z End:</b> {circuit['z_raw']}"
        )

    if circuit["city"]:
        lines.append(
            f"<b>A-End Site:</b> {circuit['city']}"
        )

    if circuit["notes"]:
        lines.append(
            f"<b>Notes:</b> {circuit['notes']}"
        )

    return "<br>".join(lines)


def make_edge_label(circuit):
    """
    Keep the visible link label relatively compact.
    """

    label = circuit["circuit"]

    if circuit["type"]:
        label += f" | {circuit['type']}"

    return label


def create_topology(circuits, output_file):
    network = Network(
        height="900px",
        width="100%",
        bgcolor="#ffffff",
        font_color="#000000",
        directed=False,
        select_menu=True,
        filter_menu=False,
    )

    nexus_devices = set()
    external_devices = set()

    for circuit in circuits:
        a_device = circuit["a_device"]
        z_device = circuit["z_device"]

        if a_device:
            if is_nexus(a_device):
                nexus_devices.add(a_device)
            else:
                external_devices.add(a_device)

        if z_device:
            if is_nexus(z_device):
                nexus_devices.add(z_device)
            else:
                external_devices.add(z_device)

    # Add Nexus switches.
    for device in sorted(nexus_devices):
        add_node(
            network,
            device,
            external=False,
        )

    # Add provider/external endpoints.
    #
    # "Unknown" needs special handling later because several circuits
    # have an Unknown Z-end. We don't want all of those circuits
    # terminating on one giant shared "Unknown" node.
    for device in sorted(external_devices):
        if device.lower() == "unknown":
            continue

        add_node(
            network,
            device,
            external=True,
        )

    for circuit in circuits:
        circuit_number = circuit["circuit"]

        a_device = circuit["a_device"]
        z_device = circuit["z_device"]

        # If one endpoint is blank, give it a circuit-specific placeholder.
        if not a_device:
            a_device = f"Unknown-A-{circuit_number}"

            add_node(
                network,
                a_device,
                label="Unknown A-End",
                external=True,
            )

        if not z_device:
            z_device = f"Unknown-Z-{circuit_number}"

            add_node(
                network,
                z_device,
                label="Unknown Z-End",
                external=True,
            )

        # Multiple rows currently say simply "Unknown".
        # Give each its own endpoint so unrelated circuits aren't joined.
        if z_device.lower() == "unknown":
            z_device = f"Unknown-Z-{circuit_number}"

            add_node(
                network,
                z_device,
                label="Unknown Z-End",
                external=True,
            )

        if a_device.lower() == "unknown":
            a_device = f"Unknown-A-{circuit_number}"

            add_node(
                network,
                a_device,
                label="Unknown A-End",
                external=True,
            )

        # Interface information is best shown as part of the hover details.
        hover = build_hover_text(circuit)

        edge_label = make_edge_label(circuit)

        # Add interface details to hover.
        if circuit["a_interface"]:
            hover += (
                f"<br><b>A Interface:</b> "
                f"{circuit['a_interface']}"
            )

        if circuit["z_interface"]:
            hover += (
                f"<br><b>Z Interface:</b> "
                f"{circuit['z_interface']}"
            )

        network.add_edge(
            a_device,
            z_device,
            label=edge_label,
            title=hover,
            width=2,
            smooth={
                "enabled": True,
                "type": "dynamic",
                "roundness": 0.2,
            },
        )

    network.set_options(
        """
        {
          "interaction": {
            "hover": true,
            "navigationButtons": true,
            "keyboard": true,
            "multiselect": true
          },

          "edges": {
            "font": {
              "size": 11,
              "align": "middle"
            },
            "smooth": {
              "enabled": true,
              "type": "dynamic"
            }
          },

          "nodes": {
            "borderWidth": 2
          },

          "physics": {
            "enabled": true,
            "solver": "barnesHut",

            "barnesHut": {
              "gravitationalConstant": -18000,
              "centralGravity": 0.2,
              "springLength": 220,
              "springConstant": 0.03,
              "damping": 0.09,
              "avoidOverlap": 0.5
            },

            "stabilization": {
              "enabled": true,
              "iterations": 1500,
              "updateInterval": 50
            }
          }
        }
        """
    )

    network.write_html(
        str(output_file),
        open_browser=False,
        notebook=False,
    )


def main():
    print(f"Reading {INPUT_FILE}")

    circuits = read_circuits(INPUT_FILE)

    print()
    print(f"Found {len(circuits)} circuits")

    nexus_devices = set()

    for circuit in circuits:
        for device in (
            circuit["a_device"],
            circuit["z_device"],
        ):
            if device and is_nexus(device):
                nexus_devices.add(device)

    print(f"Found {len(nexus_devices)} Nexus switches")

    print()
    print("Nexus devices:")

    for device in sorted(nexus_devices):
        print(f"  {device}")

    print()
    print(f"Creating {OUTPUT_FILE}")

    create_topology(
        circuits,
        OUTPUT_FILE,
    )

    print()
    print("Done.")
    print(f"Open {OUTPUT_FILE} in your browser.")


if __name__ == "__main__":
    main()